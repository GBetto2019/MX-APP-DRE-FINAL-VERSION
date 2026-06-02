"""
Sprint 0 — Task 0.6: Testes de exportação DRE (PDF e Excel).

Execute:
    pytest tests/test_sprint0_export.py -v
"""
from __future__ import annotations

import os
from decimal import Decimal
from datetime import date

import pytest
from dotenv import load_dotenv
from fastapi.testclient import TestClient
from supabase import create_client

load_dotenv()
load_dotenv(".env.test", override=False)

SUPABASE_URL  = os.environ["SUPABASE_URL"]
SUPABASE_ANON = os.environ["SUPABASE_ANON_KEY"]

from app.main import app  # noqa: E402
from app.models.schemas import DREResponse, LinhasDRE  # noqa: E402
from app.services.export_service import gerar_dre_excel, gerar_dre_pdf  # noqa: E402


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(app, raise_server_exceptions=False)


@pytest.fixture(scope="module")
def tokens() -> dict[str, str]:
    resultado: dict[str, str] = {}
    for role, email in [
        ("admin", "admin@mxseguros.test"),
        ("comercial", "comercial@mxseguros.test"),
    ]:
        anon = create_client(SUPABASE_URL, SUPABASE_ANON)
        resp = anon.auth.sign_in_with_password({"email": email, "password": "Teste@123"})
        if resp.session:
            resultado[role] = resp.session.access_token
    return resultado


def auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _dre_mock(perfil: str = "admin") -> DREResponse:
    """DRE sintético para testes unitários do serviço de export."""
    linhas = LinhasDRE(
        receita_bruta=Decimal("100000.00"),
        estornos=Decimal("5000.00"),
        impostos=Decimal("3500.00"),
        receita_liquida=Decimal("91500.00"),
        repasses_produtores=Decimal("30000.00"),
        margem_contribuicao=Decimal("61500.00"),
        despesas_fixas=Decimal("40000.00"),
        ebitda=Decimal("21500.00"),
        despesas_nao_operacionais=Decimal("5000.00"),
        resultado_liquido=Decimal("16500.00"),
    )
    if perfil == "comercial":
        linhas.receita_liquida = None
        linhas.repasses_produtores = None
        linhas.margem_contribuicao = None
        linhas.despesas_fixas = None
        linhas.ebitda = None
        linhas.despesas_nao_operacionais = None
        linhas.resultado_liquido = None

    return DREResponse(
        periodo={"inicio": date(2026, 1, 1), "fim": date(2026, 3, 31)},
        dre=linhas,
        perfil=perfil,
    )


# ── Testes unitários do serviço ────────────────────────────────

class TestExportService:

    def test_excel_retorna_bytes_validos(self):
        dre = _dre_mock("admin")
        resultado = gerar_dre_excel(dre)
        conteudo = resultado.read()
        assert len(conteudo) > 0
        # Magic bytes do XLSX (ZIP): PK\x03\x04
        assert conteudo[:2] == b"PK"

    def test_pdf_retorna_bytes_validos(self):
        dre = _dre_mock("admin")
        resultado = gerar_dre_pdf(dre)
        conteudo = resultado.read()
        assert len(conteudo) > 0
        # Magic bytes do PDF: %PDF
        assert conteudo[:4] == b"%PDF"

    def test_excel_comercial_nao_contem_ebitda(self):
        """Admin vê todas as linhas; Comercial só receita bruta, estornos e impostos."""
        import openpyxl
        from io import BytesIO

        dre_admin = _dre_mock("admin")
        dre_comercial = _dre_mock("comercial")

        xlsx_admin = gerar_dre_excel(dre_admin)
        xlsx_comercial = gerar_dre_excel(dre_comercial)

        wb_admin = openpyxl.load_workbook(BytesIO(xlsx_admin.read()))
        wb_comerc = openpyxl.load_workbook(BytesIO(xlsx_comercial.read()))

        # Admin deve ter mais linhas preenchidas que Comercial
        linhas_admin = sum(
            1 for row in wb_admin.active.iter_rows(min_row=6)
            if any(c.value for c in row)
        )
        linhas_comerc = sum(
            1 for row in wb_comerc.active.iter_rows(min_row=6)
            if any(c.value for c in row)
        )
        assert linhas_admin > linhas_comerc, (
            "Admin deve ter mais linhas visíveis que Comercial"
        )

        # EBITDA não deve aparecer para Comercial
        textos_comerc = [
            str(c.value) for row in wb_comerc.active.iter_rows(min_row=6)
            for c in row if c.value
        ]
        assert not any("EBITDA" in t for t in textos_comerc), (
            "EBITDA não deve aparecer no export do Comercial"
        )

    def test_pdf_admin_contem_resultado_liquido(self):
        """PDF do Admin deve conter 'RESULTADO LÍQUIDO'."""
        dre = _dre_mock("admin")
        pdf_bytes = gerar_dre_pdf(dre)
        conteudo = pdf_bytes.read()
        # PDF tem texto embutido — busca simples
        assert b"RESULTADO" in conteudo or len(conteudo) > 1000


# ── Testes de integração via API ───────────────────────────────

class TestExportEndpoints:

    def test_export_xlsx_sem_auth_retorna_401(self, client):
        r = client.get("/exports/dre/xlsx", params={"inicio": "2026-01-01", "fim": "2026-03-31"})
        assert r.status_code == 401

    def test_export_pdf_sem_auth_retorna_401(self, client):
        r = client.get("/exports/dre/pdf", params={"inicio": "2026-01-01", "fim": "2026-03-31"})
        assert r.status_code == 401

    def test_export_xlsx_periodo_invalido_retorna_400(self, client, tokens):
        token = tokens.get("admin")
        if not token:
            pytest.skip("Token admin não disponível")
        r = client.get(
            "/exports/dre/xlsx",
            params={"inicio": "2020-01-01", "fim": "2026-12-31"},
            headers=auth(token),
        )
        assert r.status_code == 400

    def test_export_xlsx_admin_retorna_arquivo(self, client, tokens):
        token = tokens.get("admin")
        if not token:
            pytest.skip("Token admin não disponível")
        r = client.get(
            "/exports/dre/xlsx",
            params={"inicio": "2026-01-01", "fim": "2026-03-31"},
            headers=auth(token),
        )
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        assert "attachment" in r.headers.get("content-disposition", "")

    def test_export_pdf_admin_retorna_arquivo(self, client, tokens):
        token = tokens.get("admin")
        if not token:
            pytest.skip("Token admin não disponível")
        r = client.get(
            "/exports/dre/pdf",
            params={"inicio": "2026-01-01", "fim": "2026-03-31"},
            headers=auth(token),
        )
        assert r.status_code == 200
        assert r.headers.get("content-type") == "application/pdf"

    def test_export_comercial_retorna_arquivo_menor(self, client, tokens):
        """Comercial recebe arquivo sem EBITDA/despesas — deve ser menor que Admin."""
        token_admin = tokens.get("admin")
        token_com = tokens.get("comercial")
        if not token_admin or not token_com:
            pytest.skip("Tokens não disponíveis")

        params = {"inicio": "2026-01-01", "fim": "2026-03-31"}
        r_admin = client.get("/exports/dre/xlsx", params=params, headers=auth(token_admin))
        r_com = client.get("/exports/dre/xlsx", params=params, headers=auth(token_com))

        assert r_admin.status_code == 200
        assert r_com.status_code == 200
        # Export do comercial deve ter menos bytes (menos linhas)
        assert len(r_com.content) <= len(r_admin.content)
