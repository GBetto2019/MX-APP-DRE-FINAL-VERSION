"""
Serviço de exportação do DRE para PDF e Excel.
Respeita RLS: exporta apenas o que o usuário pode ver (campos None são omitidos).
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.schemas import DREResponse


def _fmt_brl(valor: Decimal | None) -> str:
    if valor is None:
        return "—"
    sinal = "-" if valor < 0 else ""
    return f"{sinal}R$ {abs(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _pct(valor: Decimal | None, base: Decimal) -> str:
    if valor is None or base == 0:
        return "—"
    return f"{(valor / base * 100):.1f}%".replace(".", ",")


def _linhas_visiveis(dre_response: DREResponse) -> list[tuple[str, Decimal | None]]:
    d = dre_response.dre
    return [
        ("(+) RECEITA BRUTA DE COMISSÕES", d.receita_bruta),
        ("(-) Estornos e Cancelamentos",    d.estornos),
        ("(-) Impostos (Simples Nacional)",  d.impostos),
        ("(=) RECEITA LÍQUIDA",             d.receita_liquida),
        ("(-) Repasses a Produtores",        d.repasses_produtores),
        ("(=) MARGEM DE CONTRIBUIÇÃO",      d.margem_contribuicao),
        ("(-) Despesas Fixas Operacionais",  d.despesas_fixas),
        ("(=) EBITDA / LUCRO OPERACIONAL",  d.ebitda),
        ("(-) Despesas Não Operacionais",    d.despesas_nao_operacionais),
        ("(=) RESULTADO LÍQUIDO",           d.resultado_liquido),
    ]


def gerar_dre_excel(dre_response: DREResponse) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DRE"

    periodo = dre_response.periodo
    inicio = periodo.get("inicio", "")
    fim = periodo.get("fim", "")

    # Cabeçalho
    ws.merge_cells("A1:C1")
    ws["A1"] = "MX Seguros — Demonstração do Resultado do Exercício"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Período: {inicio} a {fim}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:C2")

    ws["A3"] = f"Perfil: {dre_response.perfil}"
    ws.merge_cells("A3:C3")

    # Cabeçalho da tabela
    ws["A5"] = "Linha do DRE"
    ws["B5"] = "Valor (R$)"
    ws["C5"] = "% Receita Bruta"
    for col in ("A5", "B5", "C5"):
        ws[col].font = Font(bold=True, color="FFFFFF")
        ws[col].fill = PatternFill("solid", fgColor="1F4E79")
        ws[col].alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18

    receita_bruta = dre_response.dre.receita_bruta or Decimal(0)
    linhas = _linhas_visiveis(dre_response)

    row = 6
    for label, valor in linhas:
        if valor is None:
            continue
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=float(valor)).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=_pct(valor, receita_bruta))

        # Destaque para totais
        if label.startswith("(=)"):
            for col in range(1, 4):
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="D6E4F0")
        row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_dre_pdf(dre_response: DREResponse) -> BytesIO:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []

    periodo = dre_response.periodo
    inicio = periodo.get("inicio", "")
    fim = periodo.get("fim", "")

    story.append(Paragraph(
        "<b>MX Seguros — Demonstração do Resultado do Exercício</b>",
        styles["Title"],
    ))
    story.append(Paragraph(f"Período: {inicio} a {fim}", styles["Normal"]))
    story.append(Paragraph(f"Perfil: {dre_response.perfil}", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    receita_bruta = dre_response.dre.receita_bruta or Decimal(0)
    linhas = _linhas_visiveis(dre_response)

    table_data = [["Linha do DRE", "Valor", "% Rec. Bruta"]]
    totais_idx: list[int] = []

    for i, (label, valor) in enumerate(linhas, start=1):
        if valor is None:
            continue
        table_data.append([label, _fmt_brl(valor), _pct(valor, receita_bruta)])
        if label.startswith("(=)"):
            totais_idx.append(i)

    col_widths = [10 * cm, 4.5 * cm, 3 * cm]
    table = Table(table_data, colWidths=col_widths)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for idx in totais_idx:
        style_cmds.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))
        style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#D6E4F0")))

    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    doc.build(story)
    output.seek(0)
    return output
