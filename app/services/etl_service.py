"""
Serviço de ETL via API — importação de balancete Excel (Task 3.5).

Fluxo:
1. POST /importacao/balancete  → upload do arquivo → preview (sem gravar)
2. POST /importacao/balancete/confirmar → efetiva o import em batch

Usa a lógica de categorização já existente em etl/categorizacao.py.
"""
from __future__ import annotations

import sys
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from pydantic import BaseModel
from supabase import Client

from app.logging_config import get_logger

# Importa categorizacao da pasta etl (não é package de app)
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from etl.categorizacao import classificar, SEGURADORAS  # noqa: E402

logger = get_logger(__name__)


# ── Schemas internos do ETL ───────────────────────────────────

class LancamentoPreview(BaseModel):
    linha:        int
    descricao:    str
    valor:        Decimal
    data_str:     str           # data como string (dd/mm/yyyy)
    tipo:         str           # 'despesa' | 'receita' | 'imposto' | 'revisar'
    categoria:    str | None
    subcategoria: str | None
    centro_custo: str
    mapeado:      bool          # True = classificado, False = precisa revisão
    motivo_revisar: str | None


class PreviewImportacao(BaseModel):
    total:         int
    mapeados:      int
    revisar:       int
    lancamentos:   list[LancamentoPreview]


class ResultadoImportacao(BaseModel):
    total_importado:  int
    despesas:         int
    receitas:         int
    impostos:         int
    ignorados:        int
    erros:            list[str]


# ── Parsing do Excel ──────────────────────────────────────────

def _parse_valor(val: Any) -> Decimal | None:
    """Converte célula Excel para Decimal, retorna None se inválido."""
    if val is None:
        return None
    try:
        v = float(str(val).replace(",", ".").replace("R$", "").strip())
        return Decimal(str(abs(v)))  # sempre positivo — tipo define sinal
    except (ValueError, TypeError):
        return None


def _parse_data(val: Any) -> str | None:
    """Retorna data como string YYYY-MM-DD."""
    import datetime as dt
    if val is None:
        return None
    if isinstance(val, (dt.date, dt.datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def processar_preview(arquivo_bytes: bytes) -> PreviewImportacao:
    """
    Lê o arquivo Excel e retorna preview dos lançamentos sem gravar no banco.
    Usa openpyxl para leitura.
    """
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(arquivo_bytes), data_only=True)

    lancamentos: list[LancamentoPreview] = []
    linha_global = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detectar colunas: data, descrição, valor
        header = [str(c).lower() if c else "" for c in rows[0]]
        col_data  = _detectar_coluna(header, ["data", "dt", "date", "vencimento"])
        col_desc  = _detectar_coluna(header, ["descricao", "descrição", "histórico", "historico", "nome"])
        col_valor = _detectar_coluna(header, ["valor", "debito", "crédito", "credito", "débito", "amount"])

        if col_desc is None or col_valor is None:
            continue  # aba sem estrutura reconhecível

        for row in rows[1:]:
            linha_global += 1
            if not any(row):
                continue

            descricao = str(row[col_desc]).strip() if row[col_desc] else ""
            valor     = _parse_valor(row[col_valor]) if col_valor is not None else None
            data_str  = _parse_data(row[col_data]) if col_data is not None else None

            if not descricao or valor is None or valor == 0:
                continue

            # Verifica seguradoras primeiro (evita falso-positivo ISS dentro de "comissao")
            from etl.categorizacao import _normalizar as _norm
            desc_norm = _norm(descricao)
            is_seguradora = any(seg in desc_norm for seg in SEGURADORAS)

            if is_seguradora:
                tipo, cat, subcat, cc = "receita", "", "comissao_padrao", "matriz"
            else:
                classificacao = classificar(descricao)
                tipo  = classificacao.get("tipo_lancamento", "revisar")
                cat   = classificacao.get("categoria")
                subcat = classificacao.get("subcategoria")
                cc    = classificacao.get("centro_custo", "matriz")

            motivo = None if tipo != "revisar" else "Descrição não reconhecida"

            lancamentos.append(LancamentoPreview(
                linha        = linha_global,
                descricao    = descricao[:500],
                valor        = valor,
                data_str     = data_str or "desconhecida",
                tipo         = tipo,
                categoria    = cat,
                subcategoria = subcat,
                centro_custo = cc,
                mapeado      = tipo != "revisar",
                motivo_revisar = motivo,
            ))

    mapeados = sum(1 for l in lancamentos if l.mapeado)
    return PreviewImportacao(
        total=len(lancamentos),
        mapeados=mapeados,
        revisar=len(lancamentos) - mapeados,
        lancamentos=lancamentos,
    )


def _detectar_coluna(header: list[str], candidatos: list[str]) -> int | None:
    for i, h in enumerate(header):
        for c in candidatos:
            if c in h:
                return i
    return None


# ── Efetivar import ───────────────────────────────────────────

async def efetivar_importacao(
    lancamentos: list[LancamentoPreview],
    usuario_id: str,
    db_admin: Client,
) -> ResultadoImportacao:
    """
    Insere em batch os lançamentos confirmados.
    Apenas lançamentos mapeados e não 'revisar' são inseridos.
    """
    despesas_batch: list[dict] = []
    receitas_batch: list[dict] = []
    erros: list[str] = []
    ignorados = 0

    # Buscar tipos_lancamento disponíveis para mapeamento fuzzy
    tipos_resp = db_admin.table("tipos_lancamento") \
        .select("id,categoria,natureza") \
        .eq("ativo", True).execute()
    tipos_por_cat_despesa = {
        t["categoria"]: t["id"]
        for t in (tipos_resp.data or [])
        if t["natureza"] == "despesa" and t.get("categoria")
    }
    tipos_por_nome_receita = [
        t for t in (tipos_resp.data or [])
        if t["natureza"] == "receita"
    ]
    tipo_receita_padrao = tipos_por_nome_receita[0]["id"] if tipos_por_nome_receita else None

    for lanc in lancamentos:
        if not lanc.mapeado or lanc.tipo == "revisar":
            ignorados += 1
            continue

        try:
            competencia = lanc.data_str[:7] + "-01" if len(lanc.data_str) >= 7 else date.today().replace(day=1).isoformat()

            if lanc.tipo == "despesa":
                tipo_id = tipos_por_cat_despesa.get(lanc.categoria) if lanc.categoria else None
                despesas_batch.append({
                    "subcategoria":       lanc.subcategoria or "importado",
                    "descricao":          lanc.descricao,
                    "valor":              str(lanc.valor),
                    "competencia":        competencia,
                    "centro_custo":       lanc.centro_custo,
                    "categoria":          lanc.categoria,
                    "tipo_lancamento_id": tipo_id,
                    "criado_por":         usuario_id,
                    "status":             "aprovada",
                })

            elif lanc.tipo == "receita":
                receitas_batch.append({
                    "descricao":          lanc.descricao,
                    "valor":              str(lanc.valor),
                    "competencia":        competencia,
                    "centro_custo":       lanc.centro_custo,
                    "tipo_lancamento_id": tipo_receita_padrao,
                })
            else:
                ignorados += 1

        except Exception as e:
            erros.append(f"Linha {lanc.linha}: {e}")

    # Inserir em batch
    if despesas_batch:
        db_admin.table("despesas").insert(despesas_batch).execute()

    if receitas_batch:
        db_admin.table("receitas_outras").insert(receitas_batch).execute()

    return ResultadoImportacao(
        total_importado=len(despesas_batch) + len(receitas_batch),
        despesas=len(despesas_batch),
        receitas=len(receitas_batch),
        impostos=0,   # impostos ficam na tabela impostos — not implemented aqui
        ignorados=ignorados,
        erros=erros,
    )
